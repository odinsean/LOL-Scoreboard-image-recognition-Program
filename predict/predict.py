import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import warnings
os.environ["TF_ENABLE_ONEDNN_OPTS"] = "0"
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"
warnings.filterwarnings("ignore")

import cv2
import numpy as np
import pickle
import tensorflow as tf
from openpyxl import Workbook, load_workbook
from datetime import date
from config import IMAGE_SIZE_WH, MODEL_NAME, ENCODER_PATH, RESULTS_FILE, NAME_MAP, ICON_SIZE_RATIO

ROLES = ["Top", "Jungle", "Mid", "ADC", "Support"]
RESULTS_FILE = "results.xlsx"
click_points = []

# Load Model 
print("Loading model...")
model = tf.keras.models.load_model(MODEL_NAME)

with open(ENCODER_PATH, "rb") as f:
    le = pickle.load(f)

print("Model loaded!\n")


def click_event(event, x, y, flags, param):
    """
    Handles window mouse clicks. Captures anchor pixel coordinates 
    for the Top lane and Support lane positions, rendering a 
    visual confirmation dot on the click targets.
    """
    if event == cv2.EVENT_LBUTTONDOWN:
        if len(click_points) < 2:
            click_points.append((x, y))
            cv2.circle(param, (x, y), 8, (0, 255, 0), -1)
            cv2.imshow("Screenshot", param)

            if len(click_points) == 1:
                print("Got Top icon! Now click your SUPPORT icon")
            elif len(click_points) == 2:
                print("Got Support icon! Press Q to continue")

def get_icon_positions(screenshot_path):
    """
    Launches an interactive window for the user to mark anchor roles.
    Uses the linear delta between Top and Support bounding locations
    to mathematically interpolate the centers of all five scoreboard roles.
    """
    click_points.clear()

    img = cv2.imread(screenshot_path)
    if img is None:
        print("Could not load image!")
        return None, None

    clone = img.copy()
    cv2.imshow("Screenshot", clone)
    cv2.setMouseCallback("Screenshot", click_event, clone)

    print("\nClick your TOP icon then your SUPPORT icon")
    print("Press Q when done\n")

    while True:
        key = cv2.waitKey(1) & 0xFF
        if key == ord('q') or key == ord('Q'):
            break
        if len(click_points) == 2:
            break

    cv2.destroyAllWindows()

    if len(click_points) != 2:
        print(f"Expected 2 clicks, got {len(click_points)}. Please redo.")
        return None, None

    top_x, top_y = click_points[0]
    sup_x, sup_y = click_points[1]

    # Segment pixel steps across the 4 layout gaps
    spacing_x = (sup_x - top_x) / 4
    spacing_y = (sup_y - top_y) / 4

    positions = []
    for i in range(5):
        x = round(top_x + spacing_x * i)
        y = round(top_y + spacing_y * i)
        positions.append((x, y))

    return img, positions


def predict_champions(img, positions):
    """
    Crops patches out of the canvas at designated coordinates, 
    normalizes and reshapes them to conform to tensor batch inputs,
    and runs model inference to map predictions back to champion names.
    """
    champions = []
    icon_size = max(15, round(img.shape[0] * ICON_SIZE_RATIO))

    for x, y in positions:
        crop = img[y - icon_size:y + icon_size, x - icon_size:x + icon_size]

        if crop.size == 0:
            champions.append("Unknown")
            continue

        # Pipeline Normalization & Dimensions expansion for Keras batch input
        crop = cv2.resize(crop, IMAGE_SIZE_WH)
        crop = crop.astype(np.float32) / 255.0
        crop = np.expand_dims(crop, axis=0)

        # Inference processing
        prediction = model.predict(crop, verbose=0)
        class_idx = np.argmax(prediction)
        champion = le.inverse_transform([class_idx])[0]

        # Apply name normalization maps if defined
        champion = NAME_MAP.get(champion, champion)
        champions.append(champion)

    return champions


def save_to_excel(champions):
    """
    Appends execution results to an Excel logging workbook. 
    Generates a new structured spreadsheet with headers if the file is missing,
    or increments row placements dynamically if a dataset already exists.
    """
    today = str(date.today())
    if os.path.exists(RESULTS_FILE):
        wb = load_workbook(RESULTS_FILE)
        ws = wb.active
        game_num = ws.max_row 
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["Game", "Date", "Top", "Jungle", "Mid", "ADC", "Support"])
        game_num = 1

    ws.append([game_num, today] + champions)
    wb.save(RESULTS_FILE)
    print(f"\nSaved to {RESULTS_FILE} as Game {game_num}")


# Main Loop 
print("Example inputs:")
print("  Single file:  C:\\Users\\Username\\Desktop\\game1.png")
print("  Relative:     screenshots\\game1.png")
print("  Same folder:  game1.png")
print()

while True:
    screenshot = input("\nEnter screenshot path: ").strip()

    if not os.path.exists(screenshot):
        print("File not found, please try again")
        continue

    img, positions = get_icon_positions(screenshot)

    if img is None or positions is None:
        print("Something went wrong, try again")
        continue

    champions = predict_champions(img, positions)

    print("\n=== RESULTS ===")
    for role, champ in zip(ROLES, champions):
        print(f"  {role}: {champ}")

    save_to_excel(champions)

    more = input("\nMore screenshots? (Y/N): ").strip().upper()
    if more != "Y":
        print("\nDone! Check results.xlsx for your data.")
        break
